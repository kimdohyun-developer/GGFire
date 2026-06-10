import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 1. 원본 엑셀 파일 로드
load_wb = load_workbook("/content/2026년 소집내역(수당연계, 활동실적 계).xlsx", data_only=True)
load_ws = load_wb['temp']

start_row = 6
name_col = 4      # 4열: 이름
q_col = 17        # 17열: 수당 종류 항목
amount_col = 13   # 13열: 금액

allowance_types = set()
for row_num in range(start_row, load_ws.max_row + 1):
    q_value = load_ws.cell(row=row_num, column=q_col).value
    if q_value is not None:
        allowance_types.add(str(q_value).strip())

unique_types = sorted(list(allowance_types))
classified_dict = {allowance_type: {} for allowance_type in unique_types}

for row_num in range(start_row, load_ws.max_row + 1):
    name_val = load_ws.cell(row=row_num, column=name_col).value
    q_val = load_ws.cell(row=row_num, column=q_col).value
    amount_val = load_ws.cell(row=row_num, column=amount_col).value

    if name_val is not None and q_val is not None:
        name = str(name_val).strip()
        allowance_type = str(q_val).strip()
        amount = int(amount_val) if amount_val is not None else 0

        current_allowance_dict = classified_dict[allowance_type]

        if name not in current_allowance_dict:
            current_allowance_dict[name] = amount  
        else:
            current_allowance_dict[name] += amount  

output_file_path = "/content/여비취합.xlsx"

# 스타일 공통 정의
font_title = Font(name="맑은 고딕", size=20, bold=True, color="000000")
font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF") # 흰색 글씨 헤더
font_data = Font(name="맑은 고딕", size=10)

fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # 감청색 배경
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# 🛠️ [핵심 변경] 모든 내부 칸 테두리 색상을 검은색('000000') 실선으로 변경
side_thin = Side(style='thin', color='000000')       # 내부 기본 격자선 (검은색 얇은 실선)
side_medium = Side(style='medium', color='000000')   # 표 외곽 마감선 (검은색 두꺼운 실선)

with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:

    for allowance_type in unique_types:
        name_amount_dict = classified_dict[allowance_type]

        data_list = []
        for name, total_amount in name_amount_dict.items():
            data_list.append([name, total_amount])

        df_summary = pd.DataFrame(data_list, columns=["이름", "금액"])
        df_summary = df_summary.sort_values(by="이름", ascending=True).reset_index(drop=True)
        df_summary.insert(0, "번호", range(1, len(df_summary) + 1))

        sheet_name = allowance_type[:30]
        
        # startrow=2 설정으로 3행부터 데이터가 들어가도록 고정
        df_summary.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, startcol=1)
        
        # 스타일 수정을 위해 시트 가져오기
        ws = writer.sheets[sheet_name]
        ws.sheet_view.showGridLines = True 
        
        # --------------------------------------------------
        # 디자인 시퀀스
        # --------------------------------------------------
        # 1. B, C, D열의 너비를 설정
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15

        # 2. 타이틀 (B2:D2) 지정 및 높이 설정
        ws.merge_cells("B2:D2")
        ws["B2"] = sheet_name
        ws["B2"].font = font_title
        ws["B2"].alignment = align_center
        ws.row_dimensions.height = 35 

        # 3. 전체 표 범위 설정 (2행 타이틀부터 데이터 마지막 행까지)
        start_row_idx = 2
        end_row_idx = ws.max_row

        # 4. 루프 제어를 통한 테두리 및 스타일 순차 주입
        for row_idx in range(start_row_idx, end_row_idx + 1):
            for col_idx in range(2, 5): # B열(2)부터 D열(4)까지
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # [행 위치별 맞춤형 서식 적용]
                if row_idx == 2:
                    # 2행: 타이틀 행 스타일 (병합 셀 내부 전체에 서식 균일화)
                    if col_idx == 2:
                        cell.font = font_title
                        cell.alignment = align_center
                elif row_idx == 3:
                    # 3행: 헤더 디자인 복원 (번호, 이름, 금액 덮어쓰기 및 감청색 배경)
                    headers = ["번호", "이름", "금액"]
                    cell.value = headers[col_idx - 2]
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    ws.row_dimensions.height = 25
                else:
                    # 4행 이후: 실제 데이터 행 스타일 적용
                    cell.font = font_data
                    if col_idx == 4:
                        cell.number_format = '#,##0'
                        cell.alignment = align_right
                    else:
                        cell.alignment = align_center
                    ws.row_dimensions[row_idx].height = 20

                # 🛠️ [테두리 알고리즘]
                # 1단계: 기본적으로 모든 칸에 사방 검은색 얇은 테두리(side_thin) 지정
                left_border = side_thin
                right_border = side_thin
                top_border = side_thin
                bottom_border = side_thin
                
                # 2단계: 최외곽에 해당하는 경계선만 검은색 두꺼운 테두리(side_medium)로 바꿔치기
                if col_idx == 2: left_border = side_medium        # 표의 왼쪽 끝 (B열)
                if col_idx == 4: right_border = side_medium       # 표의 오른쪽 끝 (D열)
                if row_idx == start_row_idx: top_border = side_medium   # 표의 맨 위쪽 끝 (2행)
                if row_idx == end_row_idx: bottom_border = side_medium  # 표의 맨 아래쪽 끝 (마지막 데이터행)
                
                # 완성된 테두리 객체를 셀에 할당
                cell.border = Border(
                    left=left_border, 
                    right=right_border, 
                    top=top_border, 
                    bottom=bottom_border
                )

print(f"\n[최종 성공] 스타일이 적용된 '{output_file_path}' 파일이 생성되었습니다.")

# --- 각 시트(항목)별 총 금액 및 전체 총합 출력 (기존 유지) ---
print("\n --- 각 시트(항목)별 취합 금액 보고 --- ")
grand_total_amount = 0  
grand_total_people = 0  

for allowance_type, name_dict in classified_dict.items():
    sheet_name = allowance_type[:30]
    sheet_total_amount = sum(name_dict.values())  
    sheet_people_count = len(name_dict)           

    grand_total_amount += sheet_total_amount
    grand_total_people += sheet_people_count
    print(f"▶ 시트명: [{sheet_name}] | 인원: {sheet_people_count:3d}명 | 총 금액: {sheet_total_amount:11,}원")

print("--------------------------------------------------------------------------------")
print(f"▶ 전체 항목 합계 | 인원: {grand_total_people:3d}명 | 총 금액: {grand_total_amount:11,}원")
