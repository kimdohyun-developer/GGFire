# 여비 취합 자동화
import pandas as pd
from openpyxl import load_workbook

# 1. 원본 엑셀 파일 로드
load_wb = load_workbook("/content/xsell/2026년 소집내역(수당연계, 활동실적 계).xlsx", data_only=True)
load_ws = load_wb['temp']

start_row = 6
name_col = 4      # 4열: 이름
q_col = 17        # 17열: 수당 종류 항목
amount_col = 13   # 13열: 금액
#파일형식에 따라 바꾸면 됨.

allowance_types = set()
for row_num in range(start_row, load_ws.max_row + 1):
    q_value = load_ws.cell(row=row_num, column=target_col_q).value if 'target_col_q' in locals() else load_ws.cell(row=row_num, column=q_col).value
    if q_value is not None:
        allowance_types.add(str(q_value).strip())

unique_types = sorted(list(allowance_types))

classified_dict = {allowance_type: {} for allowance_type in unique_types}


for row_num in range(start_row, load_ws.max_row + 1):
    name_val = load_ws.cell(row=row_num, column=name_col).value
    q_val = load_ws.cell(row=row_num, column=q_col).value
    amount_val = load_ws.cell(row=row_num, column=amount_col).value

    #이름과 수당 종류가 모두 있는 데이터만 처리
    if name_val is not None and q_val is not None:
        name = str(name_val).strip()
        allowance_type = str(q_val).strip()
        amount = int(amount_val) if amount_val is not None else 0

        #현재 행의 수당 종류 딕셔너리를 타겟으로 지정
        current_allowance_dict = classified_dict[allowance_type]

        #이름이 없으면 새로 추가, 있으면 금액 누적
        if name not in current_allowance_dict:
            current_allowance_dict[name] = amount  # 신규 등록
        else:
            current_allowance_dict[name] += amount  # 기존 금액에 더하기

output_file_path = "/content/xsell/여비취합.xlsx"

#print("\n=== [이름별 수당 누적 합산 및 시트 저장 시작] ===")

with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:

    for allowance_type in unique_types:
        # { '홍길동': 30000, '김철수': 15000 } 형태의 딕셔너리 가져오기
        name_amount_dict = classified_dict[allowance_type]

        # 딕셔너리를 판다스 표 구조로 변환하기 위해 리스트로 풀기
        data_list = []
        for name, total_amount in name_amount_dict.items():
            data_list.append([name, total_amount])

        # 표(DataFrame) 생성
        df_summary = pd.DataFrame(data_list, columns=["이름", "금액"])

        df_summary = df_summary.sort_values(by="이름", ascending=True).reset_index(drop=True)

        # 요구하신 "번호" 컬럼을 1번부터 순서대로 맨 앞에 추가
        df_summary.insert(0, "번호", range(1, len(df_summary) + 1))

        # 시트 저장 (글자수 31자 제한 안전장치 적용)
        sheet_name = allowance_type[:30]
        df_summary.to_excel(writer, sheet_name=sheet_name, index=False)
       # print(f"▶ '{sheet_name}' 시트: 총 {len(df_summary)}명의 누적 합산 데이터 입력 완료")

print(f"\n[최종 성공] '{output_file_path}' 파일이 생성되었습니다.")

# --------------------------------------------------
# [추가 시퀀스] 각 시트(항목)별 총 금액 및 전체 총합 출력
# --------------------------------------------------
print("\n --- 각 시트(항목)별 취합 금액 보고 --- ")

grand_total_amount = 0  # 모든 시트의 금액을 합산할 변수
grand_total_people = 0  # 모든 시트의 인원수를 합산할 변수

# 분류된 딕셔너리를 돌면서 각 항목의 총액 계산
for allowance_type, name_dict in classified_dict.items():
    sheet_name = allowance_type[:30]
    sheet_total_amount = sum(name_dict.values())  # 해당 항목의 금액 총합
    sheet_people_count = len(name_dict)           # 해당 항목의 인원수

    # 누적 전체 합계 계산
    grand_total_amount += sheet_total_amount
    grand_total_people += sheet_people_count

    # 각 시트별 결과 출력 (천 단위 콤마 추가)
    print(f"▶ 시트명: [{sheet_name}] | 인원: {sheet_people_count:3d}명 | 총 금액: {sheet_total_amount:11,}원")

print("--------------------------------------------------------------------------------")
print(f"▶ 전체 항목 합계 | 인원: {grand_total_people:3d}명 | 총 금액: {grand_total_amount:11,}원")
