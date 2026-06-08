from openpyxl import load_workbook

print("\n[작업 시작] 정리할 다른 파일을 로드합니다...")
target_file_path = "/content/xsell/2026년 개인별 의용소방대 개인 마일리지 관리서식.xlsx"
target_wb = load_workbook(target_file_path)
target_ws = target_wb['test']  # 대상 시트 지정

# --- [내 파일 구조 설정] ---
start_row = 5       # 실제 사람 이름 데이터가 시작되는 행 번호
name_col_idx = 5    # 이름이 있는 열 번호 (4열 = D열)
# ---------------------------

print("\n=== [이름 대조 및 조회 시작] ===")
success_count = 0
fail_count = 0

# 정리할 파일의 시작 행(6행)부터 데이터가 있는 마지막 줄까지 순서대로 내려갑니다.
for row_num in range(start_row, target_ws.max_row + 1):
    # 4열(D열)에서 이름 셀의 값을 읽어옵니다.
    name_value = target_ws.cell(row=row_num, column=name_col_idx).value
    
    # 만약 이름 칸이 완전히 비어있다면 데이터가 없는 빈 줄이므로 건너뜁니다.
    if name_value is None:
        continue
        
    # 정리할 파일에서 가져온 이름도 양쪽 공백을 제거하여 깨끗하게 만듭니다.
    target_name = str(name_value).strip()
    
    # [핵심 알고리즘] 요약 표 배열(summary_data)에 이 사람 이름이 존재하는지 조회
    if target_name in summary_data:
        success_count += 1
        # 대조 성공 시: 행 번호, 찾은 이름, 그리고 그 사람이 가진 활동 내역 종류를 출력
        matched_activities = list(summary_data[target_name].keys())
        print(f"▶ [{row_num}행] 매칭 성공! -> 이름: {target_name} (넣을 항목: {matched_activities})")
    else:
        fail_count += 1
        # 대조 실패 시: 요약 표에 없어서 건너뛰는 이름 출력 (인사 명부에는 있지만 이번 달 소집내역엔 없는 사람)
        print(f"❌ [{row_num}행] 매칭 실패! -> 이름: {target_name} (요약 표에 데이터 없음)")

print("\n=== [조회 및 대조 완료] ===")
print(f"총 조회한 명단: {success_count + fail_count}명")
print(f"- 요약 표와 일치하는 사람: {success_count+1}명")
print(f"- 요약 표에 없는 사람(건너뜀): {fail_count+1}명")

fixed_activity_columns = [
    "화재진압", "구조구급", "생활안전", "지원근무", "예방순찰", 
    "행사안전", "점검지원", "경계근무", "용수통로", "행사참여", 
    "급수지원", "안전교육", "캠페인활동", "정기교육", "소방학교", 
    "소방훈련", "소방기술경연대회", "불우이웃돕기", "재해복구", "기타", 
    "자격취득", "표창", "혁신제안", "안전사고", "부정사례"
]

print("\n=== [정해진 순서 배열로 데이터 입력 시작] ===")
inserted_cells_count = 0

# 5행부터 시작해서 한 줄씩 내려가며 이름 조회 및 값 입력
for row_num in range(start_row, target_ws.max_row + 1):
    name_value = target_ws.cell(row=row_num, column=name_col_idx).value
    if name_value is None:
        continue
        
    target_name = str(name_value).strip()
    
    # 요약 표 배열(summary_data)에 이 사람이 존재하는 경우
    if target_name in summary_data:
        # 이 사람의 활동 뭉치 딕셔너리 가져오기
        user_activities = summary_data[target_name]
        
        # 제공해주신 25개 항목 배열의 순서대로 칸을 한 칸씩 이동하며 대입
        for idx, activity_name in enumerate(fixed_activity_columns):
            # 이름이 5열(E열)에 있으므로, 항목들은 6열(F열)부터 1칸씩 더해가며 배치됨
            col_num = name_col_idx + 2 + idx 
            
            # 요약 데이터에 해당 항목의 값이 존재하면 입력
            if activity_name in user_activities:
                activity_value = user_activities[activity_name]
                target_ws.cell(row=row_num, column=col_num).value = activity_value
                inserted_cells_count += 1

print(f"\n=== [데이터 입력 완료] 총 {inserted_cells_count}개의 칸에 숫자를 정확히 채웠습니다. ===")

# --------------------------------------------------
# 최종 파일 저장 단계
# --------------------------------------------------
final_output_path = "/content/xsell/개인마일리지_정리완료.xlsx"
target_wb.save(final_output_path)

print(f"\n[최종 성공] 제공해주신 25개 순서대로 매칭하여 입력을 완료했습니다!")
print(f"▶ 최종 완료 파일 저장 경로: {final_output_path}")
