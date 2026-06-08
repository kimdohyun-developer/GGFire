from collections import Counter
import os
import pandas as pd
from openpyxl import load_workbook

# =========================================================================
# [1단계] 원본 자료 로딩 및 요약표 생성 (작성해주신 코드 기반)
# =========================================================================
load_wb = load_workbook("/content/xsell/2026년 소집내역(수당연계, 활동실적 계).xlsx", data_only=True)
load_ws = load_wb['temp']

col_size = load_ws.max_row 
target_col1 = 4   # 이름 열
target_col2 = 16  # 활동 내용 열
start_row = 6     

combined_data = []
for row_num in range(start_row, col_size + 1): # max_row까지만 돌도록 수정
    name_value = load_ws.cell(row=row_num, column=target_col1).value
    work_value = load_ws.cell(row=row_num, column=target_col2).value

    if name_value is not None and work_value is not None:
        combined_data.append([str(name_value).strip(), str(work_value).strip()])

activity_columns = [
    "화재진압", "구조구급", "생활안전", "지원근무", "예방순찰",
    "행사안전", "점검지원", "경계근무", "용수통로", "행사참여",
    "급수지원", "안전교육", "캠페인활동", "정기교육", "소방학교",
    "소방훈련", "소방기술경연대회", "불우이웃돕기", "재해복구", "기타",
    "자격취득", "표창", "혁신제안", "안전사고", "부정사례"
]

unique_names = sorted(list(set([row[0] for row in combined_data])))
summary_df = pd.DataFrame(0, index=unique_names, columns=activity_columns)

for name, work in combined_data:
    if work in summary_df.columns:
        summary_df.loc[name, work] += 1

# [중요] 내부 연산을 위해 0을 None으로 바꾸기 전, 복사본(딕셔너리)을 먼저 만듭니다.
# 0인 칸은 데이터 입력 단계에서 넘어가도록 처리합니다.
summary_data = summary_df.to_dict(orient="index")

# 엑셀 저장용 0 제거 후 저장
summary_df = summary_df.replace(0, None)
summary_df.to_excel("/content/xsell/소집내역_최종요약표.xlsx", index_label="이름")
print("\n[성공] 활동이 없는 칸이 모두 None 처리되어 '소집내역_최종요약표.xlsx'로 저장되었습니다.")


# =========================================================================
# [2단계] 마일리지 관리서식 파일에 데이터 매칭 입력 (요청하신 로직)
# =========================================================================
print("\n[개인 마일리지 분류 작업 시작] 관리서식 파일을 로드합니다...")
mileage_file_path = "/content/xsell/2026년 개인별 의용소방대 개인 마일리지 관리서식.xlsx"
mileage_wb = load_workbook(mileage_file_path)
mileage_ws = mileage_wb['test']  # 대상 시트 지정

# 서식 파일 구조 설정
format_start_row = 5       # 실제 사람 이름 데이터가 시작되는 행 번호
format_name_col_idx = 5    # 이름이 있는 열 번호 (5열 = E열)

print("\n=== [이름 대조 및 데이터 입력 시작] ===")
success_count = 0
fail_count = 0
inserted_cells_count = 0

# 관리서식의 5행부터 데이터가 있는 마지막 행까지 순서대로 내려갑니다.
for row_num in range(format_start_row, mileage_ws.max_row + 1):
    name_value = mileage_ws.cell(row=row_num, column=format_name_col_idx).value
    
    # 이름 칸이 비어있으면 건너뜁니다.
    if name_value is None:
        continue
        
    target_name = str(name_value).strip()
    
    # 앞서 생성한 요약 데이터(summary_data)에 이 대원의 이름이 존재하는 경우
    if target_name in summary_data:
        success_count += 1
        user_activities = summary_data[target_name]
        
        # 25개 항목 배열 순서대로 칸을 한 칸씩 이동하며 대입
        for idx, activity_name in enumerate(activity_columns):
            # 이름이 5열(E열)이므로, 항목들은 6열(F열)부터 1칸씩 더해지며 순서대로 배치됨
            col_num = format_name_col_idx + 1 + idx 
            
            # 해당 활동의 수치 가져오기
            activity_value = user_activities.get(activity_name, 0)
            
            # 값이 0보다 큰 경우에만 서식에 입력 (활동 없는 칸은 빈칸 유지)
            if activity_value > 0:
                mileage_ws.cell(row=row_num, column=col_num).value = activity_value
                inserted_cells_count += 1
    else:
        fail_count += 1

print("\n=== [조회 및 입력 완료] ===")
print(f"- 총 명단: {success_count + fail_count}명")
print(f"- 이번 달 활동 내역이 있는 대원: {success_count}명")
print(f"- 활동 내역이 없는 대원: {fail_count}명")


# --------------------------------------------------
# 최종 파일 저장 단계
# --------------------------------------------------
final_output_path = "/content/xsell/개인마일리지_정리완료.xlsx"
os.makedirs(os.path.dirname(final_output_path), exist_ok=True)
mileage_wb.save(final_output_path)

print(f"\n[최종 성공] 제공해주신 25개 순서대로 매칭하여 입력을 완료했습니다!")
print(f"▶ 최종 완료 파일 저장 경로: {final_output_path}")

