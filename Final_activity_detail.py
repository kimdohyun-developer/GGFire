import os
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 1. 파일 경로 및 시트 설정
input_file = "/content/2026년 소집내역(수당연계, 활동실적 계).xlsx"
output_file = "/content/★활동 세부내역.xlsx"
target_sheet = "temp"

# 원본 데이터 로드 (6번째 행부터 데이터 본문이므로 header=4)
df_raw = pd.read_excel(input_file, sheet_name=target_sheet, header=4)

# -------------------------------------------------------------------------
# [단계 1] 14번째 열(인덱스 13) 항목 추출 후 1차원 배열에 저장
# -------------------------------------------------------------------------
df_raw.iloc[:, 13] = df_raw.iloc[:, 13].astype(str).str.strip()
raw_unique = df_raw.iloc[:, 13].unique()

# 고유 항목들을 담을 1차원 배열(리스트) 생성 (결측치 제외)
unique_items_array = [item for item in raw_unique if item and item != "nan"]

print(
    f"📦 [1차원 배열 생성 완료] 총 {len(unique_items_array)}개의 항목이 성공적으로 저장되었습니다.\n"
)


# -------------------------------------------------------------------------
# [데이터 포맷 정제] F열 뒤 8자리 슬라이싱 및 H, I, J열 시간 복원
# -------------------------------------------------------------------------
if len(df_raw.columns) > 5:
    f_col_name = df_raw.columns[5]
    df_raw[f_col_name] = df_raw[f_col_name].astype(str).str.strip()
    df_raw[f_col_name] = df_raw[f_col_name].apply(
        lambda x: x[-8:] if (x != "nan" and len(x) >= 8) else x
    )
    df_raw[f_col_name] = df_raw[f_col_name].replace({"nan": "", "None": ""})

# 🛠️ 깨짐 버그를 막기 위해 대괄호 대신 튜플() 구조로 안전하게 변경 완료
time_cols = (7, 8, 9)
for idx in time_cols:
    if idx < len(df_raw.columns):
        col_name = df_raw.columns[idx]
        df_raw[col_name] = df_raw[col_name].astype(str).str.strip()
        df_raw[col_name] = df_raw[col_name].apply(
            lambda x: x[:5] if (":" in x and len(x) >= 8 and x != "nan") else x
        )
        df_raw[col_name] = df_raw[col_name].replace({"nan": "", "None": ""})


# -------------------------------------------------------------------------
# [단계 2] 1차원 배열 순서대로 시트를 하나씩 만들고 분류 데이터 주입
# -------------------------------------------------------------------------
# 고정 대제목 명단 (총 15개 열 = A열부터 O열까지)
fixed_headers = [
    "연번",
    "소속",
    "직위",
    "성명",
    "장소",
    "활동일자",
    "월",
    "시작시간",
    "종료시간",
    "시간",
    "합계",
    "인정시간",
    "금액",
    "비고",
    "활동내역",
]

# [열 너비 설정] 1열(A)부터 15열(O)까지 개별 고정 너비 리스트
custom_widths = [
    8,
    15,
    15,
    10,
    25,
    15,
    5,
    10,
    10,
    10,
    10,
    10,
    10,
    30,
    10,
]

os.makedirs(os.path.dirname(output_file), exist_ok=True)

print(
    "🚀 [시트 순차 생성 시작] 기본 격자 구성 후 표 가장자리에 굵은 테두리를 마감합니다..."
)

# 🎨 테두리 스타일 정의 (안쪽은 얇은 선, 바깥 가장자리는 중상 굵은 선)
thin_side = Side(style="thin", color="000000")
thick_side = Side(style="medium", color="000000")

black_border = Border(
    left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
)
center_alignment = Alignment(horizontal="center", vertical="center")

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    for index, item in enumerate(unique_items_array):

        filtered_df = df_raw[df_raw.iloc[:, 13] == item].copy()
        final_df = pd.DataFrame(columns=fixed_headers)

        for i, header_name in enumerate(fixed_headers):
            if i < len(filtered_df.columns):
                final_df[header_name] = filtered_df.iloc[:, i]

        sheet_name = f"{index + 1}. {item}"[:31]

        if sheet_name not in writer.book.sheetnames:
            ws = writer.book.create_sheet(title=sheet_name)
        else:
            ws = writer.book[sheet_name]

        # ---------------------------------------------------------------------
        # [대제목 서식] 1~2행 병합 및 보라색 서식 적용
        # ---------------------------------------------------------------------
        ws.merge_cells("A1:O2")
        title_cell = ws["A1"]
        title_cell.value = "0000년 0월 지급내역"

        title_cell.font = Font(name="맑은 고딕", size=20, bold=True, color="FFFFFF")
        title_cell.alignment = center_alignment
        title_cell.fill = PatternFill(
            start_color="6D33A6", end_color="6D33A6", fill_type="solid"
        )

        # 💡 대제목 행 높이를 각각 30으로 대폭 늘립니다 (원하는 숫자로 조절 가능)
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 28

        # 1~2행 모든 셀에 기본 격자 테두리 선입력
        for r in range(1, 3):
            for c in range(1, 16):
                ws.cell(row=r, column=c).border = black_border

        # ---------------------------------------------------------------------
        # [3행 목차 서식] 항목 제목 입력 및 노란색 스타일 적용
        # ---------------------------------------------------------------------
        ws.row_dimensions.height = 20

        yellow_fill = PatternFill(
            start_color="F2E205", end_color="F2E205", fill_type="solid"
        )
        header_font = Font(name="맑은 고딕", size=13, bold=True, color="000000")

        for col_idx, header_text in enumerate(fixed_headers):
            cell = ws.cell(row=3, column=col_idx + 1)
            cell.value = header_text
            cell.fill = yellow_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = black_border

        # ---------------------------------------------------------------------
        # [데이터 주입] 4번째 행(startrow=3)부터 실제 데이터 주입
        # ---------------------------------------------------------------------
        final_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=False,
            startrow=3,
        )

        # ---------------------------------------------------------------------
        # [본문 데이터 서식] 4행부터 데이터 끝까지 기본 격자 테두리 & 정렬
        # ---------------------------------------------------------------------
        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = 18
            for col in range(1, 16):
                data_cell = ws.cell(row=row, column=col)
                data_cell.border = black_border
                data_cell.alignment = center_alignment

        # ---------------------------------------------------------------------
        # [외곽 서식 마감] 가장자리(바깥 외곽선)만 골라내어 굵은 테두리로 덮어쓰기
        # ---------------------------------------------------------------------
        min_row, max_row = 1, ws.max_row
        min_col, max_col = 1, 15

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)

                t = thick_side if row == min_row else cell.border.top
                b = thick_side if row == max_row else cell.border.bottom
                l = thick_side if col == min_col else cell.border.left
                r = thick_side if col == max_col else cell.border.right

                cell.border = Border(top=t, bottom=b, left=l, right=r)
        # ---------------------------------------------------------------------

        # [열 너비 설정] 1열부터 15열까지 개별 지정한 너비 수동 적용
        for col_idx, width_value in enumerate(custom_widths):
            col_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[col_letter].width = width_value

        print(f"   ▶ [{index + 1}번 시트] {sheet_name} 완료 ({len(final_df)}건)")

print(
    f"\n 최종 파일 확인: {output_file}"
)
